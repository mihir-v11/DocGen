import os
import json
import mimetypes
import extract_msg
import openpyxl
import fitz  # PyMuPDF
from docx import Document
import pdfplumber
import streamlit as st
import re


def process_msg_files(msg_file_path):
    """Extracts data from a .msg file."""
    try:
        msg = extract_msg.Message(msg_file_path)
        extracted_data = f"Subject: {msg.subject}\nSender: {msg.sender}\nDate: {msg.date}\nBody: {msg.body}\n\n"
        return extracted_data
    except Exception as e:
        print(f"Error processing .msg file {msg_file_path}: {e}")
        return ""



def extract_text_from_pdf(pdf_path):
    result = "pdf_path: " + pdf_path + "\n"
    
    with fitz.open(pdf_path) as doc:
            for page in doc:
                result += page.get_text("text") + "\n"
    

    # Open PDF with pdfplumber
    with fitz.open(pdf_path) as doc:

        if doc.page_count<=10:
            
            
            table_strings = ''
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables()
                    for table in tables:
                        if table:
                            table_str = ""
                            for row in table:
                                row_str = "\t".join(cell if cell is not None else "" for cell in row)
                                table_str += row_str + "\n"
                            table_strings +=table_str.strip()
            result += "Table data: " + table_strings

    return result
def extract_text_from_docx(docx_path):
    """Extracts text from a DOCX file."""
    try:
        
        doc = Document(docx_path)
        result = ""
        for block in doc.element.body:
            if block.tag.endswith('p'):
                # Handle paragraphs manually (like in your original logic)
                paragraph = block
                text = ''.join(node.text for node in paragraph.xpath('.//w:t') if node.text)
                if text.strip():
                    result += text.strip() + "\n"

            elif block.tag.endswith('tbl'):
                # Convert XML table to docx.table.Table
                for tbl in doc.tables:
                    if tbl._element == block:
                        table_data = []
                        for row in tbl.rows:
                            row_data = {}
                            for i, cell in enumerate(row.cells):
                                row_data[f"col_{i}"] = cell.text.strip()
                            table_data.append(row_data)

                        table_json = json.dumps(table_data, ensure_ascii=False)
                        result += table_json + "\n"
                        break  # Only match once to avoid duplicates

        return result.strip()

    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return None


def extract_text_from_xlsx(xlsx_path):
    """Extracts text from an XLSX file."""
    text = ""
    try:
        workbook = openpyxl.load_workbook(xlsx_path)
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        text += str(cell.value) + "\n"
    except Exception as e:
        print(f"Error processing XLSX file {xlsx_path}: {e}")
    return text.strip() + "\n"

def extract_text_from_txt(txt_path):
    """Extracts text from a TXT file."""
    text = ""
    try:
        with open(txt_path, "r", encoding="utf-8") as f:
            text = f.read()
    except UnicodeDecodeError:
        print(f"Could not decode {txt_path} with utf-8, trying default encoding")
        try:
            with open(txt_path, "r") as f:
                text = f.read()
        except Exception as e:
            print(f"Error processing TXT file {txt_path}: {e}")
    return text.strip() + "\n"

def extract_data_from_file(file_path):
    """Extracts data from a single file."""
    print(f"Processing file: {file_path}")
    mime_type, _ = mimetypes.guess_type(file_path)
    extracted_data = ""
    
    if mime_type == "application/pdf":
        extracted_data = extract_text_from_pdf(file_path)
    # elif mime_type in ["application/vnd.openxmlformats-officedocument.wordprocessingml.document", "application/msword"]:
    #     if file_path.endswith(".doc"):
    #         docx_path = convert_doc_to_docx(file_path)
    #         if docx_path:
    #             extracted_data = extract_text_from_docx(docx_path)
    #     else:
    #         extracted_data = extract_text_from_docx(file_path)
    elif mime_type == "text/plain":
        extracted_data = extract_text_from_txt(file_path)
    elif mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        extracted_data = extract_text_from_xlsx(file_path)
    elif file_path.endswith(".msg"):
        extracted_data = process_msg_files(file_path)
    else:
        print(f"Unsupported file type: {mime_type} for file {file_path}")
        
    return extracted_data

def extract_data_from_folder(folder):
    """Extracts data from .pdf, .docx, .txt, .xlsx, and .msg files within a folder and its subfolders."""
    extracted_data = ""
    for root, _, files in os.walk(folder):
        for filename in files:
            file_path = os.path.join(root, filename)
            extracted_data += extract_data_from_file(file_path)
    return extracted_data

def data_extraction(input_files_path):
    """Extracts data from multiple input file paths (files or folders) and returns as a string."""
    all_extracted_data = ""
    
    for path in input_files_path:

        print(f"Processing path: {path}")
        if not os.path.exists(path):
            print(f"Path does not exist: {path}")
        elif os.path.isfile(path):
            # all_extracted_data += "pdf_file_path: " + path + "\n"
            all_extracted_data += extract_data_from_file(path)
        elif os.path.isdir(path):
            all_extracted_data += extract_data_from_folder(path)
    
    return all_extracted_data







